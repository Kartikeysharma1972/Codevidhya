"""Parse the CBSE Grade-wise TOC Excel into a JSON file consumable by the frontend."""
import json
import re
import sys
from pathlib import Path

import openpyxl

XLSX_PATH = Path(__file__).parent / "GRADE Wise TOC.xlsx"
OUT_PATH = Path(__file__).parent.parent / "frontend" / "src" / "data" / "cbseTOC.json"


def clean(text):
    if text is None:
        return ""
    s = str(text).strip()
    # normalise mangled unicode (curly quotes etc.) common in copy-pasted CBSE docs
    s = s.replace("﻿", "").replace("’", "'").replace("‘", "'")
    s = s.replace("“", '"').replace("”", '"').replace("–", "-").replace("—", "-")
    s = re.sub(r"\s+", " ", s)
    return s


def parse():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        sh = wb[sheet_name]
        m = re.search(r"(\d+)", sheet_name)
        if not m:
            continue
        grade_key = f"Grade {int(m.group(1))}"
        rows = list(sh.iter_rows(values_only=True))
        if not rows:
            continue
        # find header row and locate column positions
        header_idx = 0
        col_subject = 0
        col_chapter_no = 1
        col_chapter_name = 2
        col_description = 3
        for i, r in enumerate(rows[:5]):
            cells = [clean(c).lower() if c is not None else "" for c in r]
            if any("subject" in c for c in cells if c) or any("chapter" in c for c in cells if c):
                header_idx = i
                for idx, c in enumerate(cells):
                    if "subject" in c:
                        col_subject = idx
                    elif "chapter" in c and "no" in c:
                        col_chapter_no = idx
                    elif ("chapter" in c and "name" in c) or ("topic name" in c):
                        col_chapter_name = idx
                    elif "description" in c or "concept" in c:
                        col_description = idx
                break
        current_subject = None
        subjects = {}
        for row in rows[header_idx + 1:]:
            if not row:
                continue
            def cell(i):
                return clean(row[i]) if i < len(row) else ""
            subject_cell = cell(col_subject)
            chapter_no = cell(col_chapter_no)
            chapter_name = cell(col_chapter_name)
            description = cell(col_description)
            if subject_cell:
                current_subject = subject_cell.upper().strip()
                subjects.setdefault(current_subject, [])
            if not current_subject:
                continue
            if not chapter_name and not description:
                continue
            topic = chapter_name or description.split(":")[0]
            if not topic:
                continue
            subjects[current_subject].append({
                "chapter": chapter_no,
                "topic": topic,
                "description": description,
            })
        # drop empty subjects
        subjects = {k: v for k, v in subjects.items() if v}
        if subjects:
            result[grade_key] = subjects
    return result


def normalise_subject_names(data):
    """Map abbreviations like MATHS -> Mathematics to Title Case consistent names."""
    mapping = {
        "MATHS": "Mathematics",
        "MATHEMATICS": "Mathematics",
        "EVS": "Environmental Studies",
        "ENVIRONMENTAL STUDIES": "Environmental Studies",
        "ENGLISH": "English",
        "ENGLISH LIT": "English Literature",
        "ENGLISH LITERATURE": "English Literature",
        "ENGLISH GRAM": "English Grammar",
        "ENGLISH GRAMMAR": "English Grammar",
        "ENGLISH CORE": "English",
        "HINDI": "Hindi",
        "HINDI LIT": "Hindi Literature",
        "HINDI LITERATURE": "Hindi Literature",
        "HINDI GRAM": "Hindi Grammar",
        "HINDI GRAMMAR": "Hindi Grammar",
        "HINDI CORE": "Hindi",
        "SCIENCE": "Science",
        "PHYSICS": "Physics",
        "CHEMISTRY": "Chemistry",
        "BIOLOGY": "Biology",
        "SST": "Social Science",
        "SOCIAL SCIENCE": "Social Science",
        "SOCIAL STUDIES": "Social Science",
        "SST (HIST)": "History",
        "SST (GEOG)": "Geography",
        "SST (POL)": "Political Science",
        "SST (ECON)": "Economics",
        "COMPUTER SCIENCE": "Computer Science",
        "INFORMATICS PRACTICES": "Informatics Practices",
        "HISTORY": "History",
        "GEOGRAPHY": "Geography",
        "POLITICAL SCIENCE": "Political Science",
        "ECONOMICS": "Economics",
        "ACCOUNTANCY": "Accountancy",
        "BUSINESS ST": "Business Studies",
        "BUSINESS STUDIES": "Business Studies",
        "PSYCHOLOGY": "Psychology",
        "SOCIOLOGY": "Sociology",
        "PHYSICAL EDUCATION": "Physical Education",
        "SANSKRIT": "Sanskrit",
    }
    out = {}
    for grade, subs in data.items():
        out[grade] = {}
        for sname, topics in subs.items():
            key = sname.upper().strip()
            pretty = mapping.get(key, sname.title())
            # merge if already exists
            existing = out[grade].get(pretty, [])
            existing.extend(topics)
            out[grade][pretty] = existing
    return out


def main():
    data = parse()
    data = normalise_subject_names(data)
    # dump summary to stdout (ASCII-safe)
    summary = {g: {s: len(t) for s, t in subs.items()} for g, subs in data.items()}
    sys.stdout.write(json.dumps(summary, indent=2, ensure_ascii=True) + "\n")
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    sys.stdout.write(f"Wrote {OUT_PATH}\n")


if __name__ == "__main__":
    main()
